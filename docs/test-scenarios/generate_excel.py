"""
Lucky Piggy 추첨 프로세스 QA 테스트 시나리오 Excel 생성 스크립트
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 색상 팔레트 ───────────────────────────────────────────
COLORS = {
    "header_bg":   "1E3A5F",   # 진네이비
    "header_fg":   "FFFFFF",
    "group0":      "D9E1F2",   # 사전준비 - 연파랑
    "group1":      "E2EFDA",   # 응모 - 연초록
    "group2":      "FFF2CC",   # 응모마감 - 연노랑
    "group3":      "FCE4D6",   # 당첨자판정 - 연주황
    "group4":      "EAD1DC",   # 당첨자발표 - 연분홍
    "group5":      "D9EAD3",   # 계좌제출 - 민트
    "group6":      "CFE2F3",   # 어드민관리 - 하늘
    "group7":      "F4CCCC",   # 수동당첨자 - 연빨강
    "group8":      "FFE599",   # 만료처리 - 연금
    "group9":      "D5E8D4",   # 초기화 - 연녹
    "groupE":      "EEEEEE",   # 예외케이스 - 연회색
    "row_odd":     "FFFFFF",
    "row_even":    "F8F9FA",
    "pass_bg":     "D4EDDA",
    "fail_bg":     "F8D7DA",
    "skip_bg":     "FFF3CD",
}

def thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def header_font():
    return Font(name="맑은 고딕", bold=True, color=COLORS["header_fg"], size=10)

def group_font(bold=False):
    return Font(name="맑은 고딕", bold=bold, size=9)

def apply_header(ws, row, cols, bg_color):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = header_font()
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()

def write_row(ws, row_num, values, bg_color, bold=False):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font = Font(name="맑은 고딕", bold=bold, size=9)
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                 horizontal="center" if col in (1, 7) else "left")
        c.border = thin_border()

# ══════════════════════════════════════════════════════════
# Sheet 1: 전체 목차
# ══════════════════════════════════════════════════════════
ws_toc = wb.active
ws_toc.title = "목차"
ws_toc.sheet_view.showGridLines = False

ws_toc["B2"] = "Lucky Piggy 추첨 프로세스 QA 테스트 시나리오"
ws_toc["B2"].font = Font(name="맑은 고딕", bold=True, size=16, color="1E3A5F")
ws_toc["B3"] = "2026-05-19 | 스테이징 + 프로덕션 테스트용"
ws_toc["B3"].font = Font(name="맑은 고딕", size=10, color="666666")

ws_toc.row_dimensions[2].height = 30
ws_toc.row_dimensions[3].height = 16

toc_header = ["그룹", "시트명", "테스트 항목 수", "설명"]
apply_header(ws_toc, 5, toc_header, COLORS["header_bg"])

toc_data = [
    ("G-0",  "G0_사전준비",     5,  "테스트 환경 세팅, 회차·유저·티켓 초기 확인"),
    ("G-1",  "G1_응모",         12, "정상 응모, 티켓 차감, 유효성 오류 케이스"),
    ("G-2",  "G2_응모마감",     7,  "close-draw 시뮬레이션 (어드민 패널 Step1)"),
    ("G-3",  "G3_당첨자판정",   14, "judge_draw_winners RPC, 등수별 판정, UI 가드"),
    ("G-4",  "G4_당첨자발표",   7,  "publish-and-open 시뮬레이션 (어드민 패널 Step3)"),
    ("G-5",  "G5_계좌제출",     14, "submit-bank-account, 유효성, 마감 초과"),
    ("G-6",  "G6_어드민관리",   15, "어드민 당첨자 목록, 확인 토글, 지급 처리, 공개 RPC"),
    ("G-7",  "G7_수동당첨자",   8,  "AddWinnerDialog, 초대코드 생성, 삭제 보호"),
    ("G-8",  "G8_만료처리",     7,  "expire_draw_winners cron, 조건별 제외"),
    ("G-9",  "G9_초기화",       5,  "TestControlPanel 초기화, 상태 롤백 검증"),
    ("G-E",  "GE_예외엣지케이스", 20, "Race condition, FK, RLS, DB CHECK 제약"),
]

group_colors = [
    COLORS["group0"], COLORS["group1"], COLORS["group2"], COLORS["group3"],
    COLORS["group4"], COLORS["group5"], COLORS["group6"], COLORS["group7"],
    COLORS["group8"], COLORS["group9"], COLORS["groupE"],
]

for i, (grp, sheet, cnt, desc) in enumerate(toc_data):
    r = 6 + i
    bg = group_colors[i]
    values = [grp, sheet, cnt, desc]
    write_row(ws_toc, r, values, bg)
    ws_toc.row_dimensions[r].height = 18

ws_toc.column_dimensions["A"].width = 3
ws_toc.column_dimensions["B"].width = 8
ws_toc.column_dimensions["C"].width = 20
ws_toc.column_dimensions["D"].width = 10
ws_toc.column_dimensions["E"].width = 50

# 합계 행
total_row = 6 + len(toc_data)
ws_toc.cell(row=total_row, column=2, value="합계").font = Font(name="맑은 고딕", bold=True, size=9)
ws_toc.cell(row=total_row, column=3, value=sum(x[2] for x in toc_data)).font = Font(name="맑은 고딕", bold=True, size=9)
ws_toc.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor=COLORS["header_bg"])
ws_toc.cell(row=total_row, column=3).fill = PatternFill("solid", fgColor=COLORS["header_bg"])
ws_toc.cell(row=total_row, column=2).font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)
ws_toc.cell(row=total_row, column=3).font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)
ws_toc.cell(row=total_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
ws_toc.cell(row=total_row, column=3).alignment = Alignment(horizontal="center", vertical="center")

# ══════════════════════════════════════════════════════════
# 시나리오 데이터
# ══════════════════════════════════════════════════════════
COLS = ["번호", "테스트 항목", "사전 조건", "실행 방법", "기대 결과", "DB 확인 쿼리", "비고", "결과", "메모"]
COL_WIDTHS = [8, 35, 30, 38, 35, 50, 28, 8, 20]

GROUPS = [
    {
        "title": "G0_사전준비",
        "color": COLORS["group0"],
        "desc": "사전 준비 — 회차 초기화 / 테스트 데이터",
        "rows": [
            ("T-000", "테스트 회차 active 상태 확인",
             "DB에 active 회차가 존재",
             "SELECT id, round_number, status FROM draws WHERE status='active' LIMIT 1; 실행",
             "1개 회차 반환, status='active'",
             "SELECT id, round_number, status, start_date, end_date, draw_date FROM draws ORDER BY round_number DESC LIMIT 5;",
             "없으면 T-001로 생성"),
            ("T-001", "테스트 회차 신규 생성",
             "DB에 active 회차가 없을 때",
             "어드민 또는 직접 INSERT: round_number=N, status='active', start_date=now(), end_date=now()+7d",
             "draws 테이블에 신규 행 생성",
             "SELECT id, round_number, status FROM draws WHERE round_number = N;",
             "end_date는 현재 시각 이후로 설정"),
            ("T-002", "테스트용 유저 응모권 잔액 확인",
             "테스트 유저 존재",
             "SELECT user_id, total, spent, total-spent AS available FROM ticket_balances WHERE user_id='<uid>';",
             "available >= 1",
             "SELECT total, spent FROM ticket_balances WHERE user_id='<uid>';",
             "없으면 earn-tickets로 충전"),
            ("T-003", "draw_prizes 설정 확인",
             "대상 draws 존재",
             "SELECT prize_rank, name, amount FROM draw_prizes WHERE draw_id='<draw_id>' ORDER BY prize_rank;",
             "1~5등 상품 정의 존재 (없어도 수동 당첨자 추가 시 자동 생성)",
             "SELECT prize_rank, amount FROM draw_prizes WHERE draw_id='<id>' ORDER BY prize_rank;",
             "judge_draw_winners 실행 후 prize_id 연결 필수"),
            ("T-004", "feature_flag draw_entry 활성화 확인",
             "feature_flags 테이블 존재",
             "SELECT flag_name, is_enabled FROM feature_flags WHERE flag_name='draw_entry';",
             "is_enabled=true",
             "SELECT flag_name, is_enabled FROM feature_flags;",
             "false이면 enter-draw 응모 차단됨"),
        ]
    },
    {
        "title": "G1_응모",
        "color": COLORS["group1"],
        "desc": "Step 1 — 응모 (enter-draw RPC)",
        "rows": [
            ("T-101", "정상 응모 — 1장",
             "draws.status='active', 유저 티켓 잔액 >= 1",
             "enter-draw RPC 호출: {draw_id, user_id, entries:[[1,7,14,23,38,43]]}",
             "응답 {total_count:1, entries:[...]}, draw_entries 1행 생성",
             "SELECT id, lottery_numbers, status FROM draw_entries WHERE draw_id='<id>' AND user_id='<uid>' ORDER BY entered_at DESC LIMIT 1;",
             "tickets_used=1"),
            ("T-102", "정상 응모 — 여러 장",
             "티켓 잔액 >= 3",
             "entries에 3개 배열 전달",
             "draw_entries 3행 생성, ticket_balances.spent +3",
             "SELECT count(*) FROM draw_entries WHERE draw_id='<id>' AND user_id='<uid>';",
             "세트 간 번호 중복 허용"),
            ("T-103", "응모권 차감 검증",
             "응모 전 spent=S",
             "T-101 실행 후 ticket_balances 재조회",
             "ticket_balances.spent = S+1",
             "SELECT total, spent FROM ticket_balances WHERE user_id='<uid>';",
             ""),
            ("T-104", "ticket_transactions 기록 검증",
             "T-101 실행 후",
             "SELECT type, amount, reference_id FROM ticket_transactions WHERE user_id='<uid>' ORDER BY created_at DESC LIMIT 1;",
             "type='draw_entry', amount=-N, reference_id=draw_id",
             "SELECT type, amount, reference_id FROM ticket_transactions WHERE user_id='<uid>' ORDER BY created_at DESC LIMIT 3;",
             "단일 합계 행"),
            ("T-105", "draws.total_entries 집계 검증",
             "T-101 실행 후",
             "SELECT total_entries FROM draws WHERE id='<draw_id>';",
             "이전보다 +N 증가",
             "SELECT id, total_entries FROM draws WHERE id='<draw_id>';",
             "모니터링용 컬럼"),
            ("T-106", "잔액 부족 시 응모 거부",
             "티켓 잔액=0",
             "enter-draw 호출",
             "EXCEPTION 'insufficient_tickets', draw_entries 미생성",
             "SELECT total-spent FROM ticket_balances WHERE user_id='<uid>';",
             "롤백 확인"),
            ("T-107", "inactive 회차에 응모 거부",
             "draws.status='closed'",
             "enter-draw 호출",
             "EXCEPTION 'draw_not_active'",
             "SELECT status FROM draws WHERE id='<draw_id>';",
             "upcoming/drawn/completed도 동일"),
            ("T-108", "번호 범위 오류 (0 또는 46 포함)",
             "active 회차",
             "entries에 [0,7,14,23,38,43] 전달",
             "EXCEPTION 'invalid_entry_range'",
             "—",
             "1~45 범위 외 거부"),
            ("T-109", "번호 개수 오류 (5개 또는 7개)",
             "active 회차",
             "entries에 [1,7,14,23,38] 전달 (5개)",
             "EXCEPTION 'invalid_entry_length'",
             "—",
             "정확히 6개 요구"),
            ("T-110", "한 세트 내 중복 번호",
             "active 회차",
             "entries에 [1,1,14,23,38,43] 전달",
             "EXCEPTION 'duplicate_numbers_in_entry'",
             "—",
             "세트 간 중복은 허용"),
            ("T-111", "빈 entries 배열",
             "active 회차",
             "entries에 [] 전달",
             "EXCEPTION 'empty_entries'",
             "—",
             ""),
            ("T-112", "ticket_balances 행 없는 유저",
             "신규 유저 (잔액 행 미생성)",
             "enter-draw 호출",
             "EXCEPTION 'ticket_balances_not_found'",
             "SELECT * FROM ticket_balances WHERE user_id='<uid>';",
             ""),
        ]
    },
    {
        "title": "G2_응모마감",
        "color": COLORS["group2"],
        "desc": "Step 2 — 응모 마감 (close-draw)",
        "rows": [
            ("T-201", "어드민 패널 Step1 클릭 — 정상 마감",
             "draws.status='active'",
             "TestControlPanel > [Step 1] 응모 마감 클릭",
             "draws.status='closed', UI 배지 '응모 마감'으로 변경",
             "SELECT status FROM draws WHERE id='<draw_id>';",
             "closeDrawForTest: .eq('status','active') 조건"),
            ("T-202", "이미 closed 회차 재마감 시도",
             "draws.status='closed'",
             "Step1 버튼 상태 확인",
             "버튼 disabled, 클릭 불가 (step1Active=false)",
             "—",
             ""),
            ("T-203", "close-draw EF 직접 호출 — active+end_date 초과",
             "status='active', end_date <= now()",
             "POST /functions/v1/close-draw (CRON_SECRET 헤더)",
             "{closed:1, upcomingCreated:1}, draws.status='closed', 다음 회차 upcoming 생성",
             "SELECT round_number, status FROM draws ORDER BY round_number DESC LIMIT 3;",
             "end_date 미초과면 closed 전환 안 됨"),
            ("T-204", "close-draw — 다음 회차 upcoming 생성 멱등성",
             "같은 회차로 close-draw 두 번 호출",
             "동일 요청 2회 전송",
             "두 번째 호출 시 upcomingCreated:0 (중복 스킵)",
             "SELECT count(*) FROM draws WHERE round_number=N+1;",
             "ON CONFLICT / maybeSingle로 중복 방지"),
            ("T-205", "close-draw — 다음 회차 날짜 계산 검증",
             "토요일 draw_date 기준",
             "생성된 upcoming 회차 날짜 확인",
             "start_date=다음 주 일요일, end_date=다음 주 토요일 20:00, draw_date=다음 주 토요일",
             "SELECT start_date, end_date, draw_date FROM draws WHERE status='upcoming';",
             "nextDrawStartKst(baseDate) 기준"),
            ("T-206", "응모 마감 후 신규 응모 불가",
             "draws.status='closed'",
             "enter-draw 호출",
             "EXCEPTION 'draw_not_active'",
             "SELECT status FROM draws WHERE id='<draw_id>';",
             ""),
            ("T-207", "close-draw 인증 실패 (잘못된 토큰)",
             "—",
             "Authorization 헤더 없이 POST",
             "403 Forbidden",
             "—",
             "CRON_SECRET 불일치"),
        ]
    },
    {
        "title": "G3_당첨자판정",
        "color": COLORS["group3"],
        "desc": "Step 3 — 당첨자 판정 (judge_draw_winners RPC)",
        "rows": [
            ("T-301", "어드민 패널 Step2 — 번호 입력 후 판정",
             "draws.status='closed'",
             "Step2 클릭 → 당첨번호 6개 + 보너스 1개 입력 → 판정 실행",
             "draws.winning_numbers 저장, judge_draw_winners RPC 실행, draws.status='drawn', draw_winners 자동 생성",
             "SELECT winning_numbers, bonus_number, status FROM draws WHERE id='<draw_id>';",
             ""),
            ("T-302", "1등 판정 — 6개 모두 일치",
             "응모자 lottery_numbers=[1,7,14,23,38,43], 당첨번호=[1,7,14,23,38,43]",
             "T-301 실행",
             "draw_entries.status='won', draw_winners.prize_rank=1, match_count=6",
             "SELECT match_count, prize_rank FROM draw_winners WHERE draw_id='<id>';",
             ""),
            ("T-303", "2등 판정 — 5개 일치 + 보너스 일치",
             "lottery_numbers 5개 일치 + 보너스번호 포함",
             "T-301 실행",
             "prize_rank=2, match_count=5, has_bonus=true",
             "SELECT prize_rank, match_count FROM draw_winners WHERE draw_entry_id='<entry_id>';",
             ""),
            ("T-304", "3등 판정 — 5개 일치, 보너스 불일치",
             "lottery_numbers 5개 일치, 보너스 불일치",
             "T-301 실행",
             "prize_rank=3, match_count=5",
             "SELECT prize_rank, match_count FROM draw_winners WHERE draw_entry_id='<entry_id>';",
             ""),
            ("T-305", "4등 판정 — 4개 일치",
             "lottery_numbers 4개 일치",
             "T-301 실행",
             "prize_rank=4, match_count=4",
             "SELECT prize_rank, match_count FROM draw_winners WHERE draw_entry_id='<entry_id>';",
             ""),
            ("T-306", "5등 판정 — 3개 일치",
             "lottery_numbers 3개 일치",
             "T-301 실행",
             "prize_rank=5, match_count=3",
             "SELECT prize_rank, match_count FROM draw_winners WHERE draw_entry_id='<entry_id>';",
             ""),
            ("T-307", "미당첨 — 2개 이하 일치",
             "lottery_numbers 2개 이하 일치",
             "T-301 실행",
             "draw_entries.status='lost', draw_winners 미생성",
             "SELECT status FROM draw_entries WHERE draw_id='<id>' AND user_id='<uid>';",
             "prize_rank=null"),
            ("T-308", "judge_draw_winners 멱등성 확인",
             "T-301 실행 완료 후",
             "동일 draw_id로 RPC 재호출",
             "draw_winners 중복 생성 없음 (ON CONFLICT DO NOTHING)",
             "SELECT count(*) FROM draw_winners WHERE draw_id='<id>';",
             "draw_entry_id UNIQUE 보장"),
            ("T-309", "winning_numbers 없는 상태에서 RPC 직접 호출",
             "draws.winning_numbers=null",
             "winning_numbers 설정 없이 judge_draw_winners RPC 호출",
             "EXCEPTION 'winning_numbers_not_set'",
             "SELECT winning_numbers FROM draws WHERE id='<draw_id>';",
             ""),
            ("T-310", "번호 6개 미만 입력 시 UI 차단",
             "Step2 폼 표시 중",
             "번호 6개 중 1개 비우고 판정 실행 클릭",
             "에러 메시지 '당첨번호 6개와 보너스번호를 1~45 사이로 입력하세요.' 표시",
             "—",
             "클라이언트 유효성 검사"),
            ("T-311", "범위 외 번호 입력 시 UI 차단",
             "Step2 폼 표시 중",
             "번호에 0 또는 46 입력 후 판정 실행",
             "에러 메시지 표시, RPC 미호출",
             "—",
             "클라이언트 유효성 검사"),
            ("T-312", "prize_id 연결 검증 (draw_prizes 있을 때)",
             "draw_prizes에 해당 draw_id 등수 설정됨",
             "T-301 실행",
             "draw_winners.prize_id IS NOT NULL",
             "SELECT dw.prize_rank, dw.prize_id, dp.amount FROM draw_winners dw LEFT JOIN draw_prizes dp ON dp.id=dw.prize_id WHERE dw.draw_id='<id>';",
             ""),
            ("T-313", "prize_id 연결 검증 (draw_prizes 없을 때)",
             "draw_prizes 미설정",
             "T-301 실행",
             "draw_winners.prize_id IS NULL (오류 없이 진행)",
             "SELECT prize_id FROM draw_winners WHERE draw_id='<id>';",
             ""),
            ("T-314", "집계 응답 검증",
             "T-301 실행",
             "RPC 응답 확인",
             "{total: N, won: W, lost: L, winners_inserted: W_count} 모두 정확",
             "SELECT count(*) FILTER (WHERE status='won'), count(*) FILTER (WHERE status='lost') FROM draw_entries WHERE draw_id='<id>';",
             ""),
        ]
    },
    {
        "title": "G4_당첨자발표",
        "color": COLORS["group4"],
        "desc": "Step 4 — 당첨자 발표 (publish-and-open)",
        "rows": [
            ("T-401", "어드민 패널 Step3 — 당첨자 발표",
             "draws.status='drawn', 다음 회차 upcoming 존재",
             "Step3 클릭 → 확인 다이얼로그 OK",
             "현재 회차 status='completed', 다음 회차 status='active'",
             "SELECT round_number, status FROM draws ORDER BY round_number DESC LIMIT 3;",
             ""),
            ("T-402", "발표 후 다음 회차 upcoming→active 전환",
             "upcoming 회차 존재 (round+1)",
             "T-401 실행",
             "round_number=current+1 행의 status='active'",
             "SELECT round_number, status FROM draws WHERE round_number=N+1;",
             "publishDrawForTest 2단계"),
            ("T-403", "다음 회차 없을 때 새 active 회차 자동 생성",
             "upcoming 회차 없음",
             "publishDrawForTest 실행",
             "round_number=current+1 신규 active 회차 INSERT",
             "SELECT round_number, status, start_date FROM draws ORDER BY round_number DESC LIMIT 1;",
             "폴백 로직"),
            ("T-404", "publish-and-open EF 직접 호출 — drawn→completed",
             "status='drawn' 회차 존재",
             "POST /functions/v1/publish-and-open (CRON_SECRET)",
             "{completed:1, pushed:[...], activated:1}",
             "SELECT status FROM draws WHERE id='<draw_id>';",
             "실제 cron 동작 시뮬레이션"),
            ("T-405", "당첨자 푸시 알림 발송 확인",
             "status='drawn' 회차, draw_entries.status='won' 당첨자 존재",
             "T-404 실행",
             "send-push invoke 호출, 당첨자 user_id 목록 전달 (best-effort)",
             "SELECT user_id FROM draw_entries WHERE draw_id='<id>' AND status='won';",
             "실패해도 완료 처리 진행"),
            ("T-406", "당첨자 없는 회차 발표",
             "draw_entries 전체 status='lost'",
             "T-404 실행",
             "completed 전환, pushed=[{drawId, pushed:0}]",
             "SELECT count(*) FROM draw_winners WHERE draw_id='<id>';",
             "정상 처리"),
            ("T-407", "발표 전(drawn 이전) Step3 비활성화 확인",
             "status='closed' 또는 'active'",
             "어드민 패널 Step3 버튼 상태 확인",
             "버튼 disabled (step3Active = status==='drawn')",
             "—",
             "UI 가드"),
        ]
    },
    {
        "title": "G5_계좌제출",
        "color": COLORS["group5"],
        "desc": "당첨자 계좌 제출 — submit-bank-account",
        "rows": [
            ("T-501", "정상 계좌 제출",
             "draw_winners.user_id=uid, payment_status='pending', account_submitted_at=null, deadline 7일 이내",
             "submit-bank-account POST: {draw_id, real_name, bank_name, account_number, account_holder}",
             "{success:true}, draw_winners 업데이트 (account_submitted_at 설정)",
             "SELECT account_submitted_at, bank_name FROM draw_winners WHERE draw_id='<id>' AND user_id='<uid>';",
             "real_name 로그 출력 없음"),
            ("T-502", "winner_comment 포함 제출",
             "T-501 조건 + winner_comment 50자 이내",
             "body에 winner_comment 포함",
             "draw_winners.winner_comment 저장",
             "SELECT winner_comment FROM draw_winners WHERE draw_id='<id>' AND user_id='<uid>';",
             ""),
            ("T-503", "winner_comment 50자 초과 거부",
             "T-501 조건",
             "winner_comment 51자 전달",
             "{success:false, error:'invalid_comment'}",
             "—",
             ""),
            ("T-504", "real_name 형식 오류 (한글 아님)",
             "T-501 조건",
             "real_name='John' 전달",
             "{success:false, error:'invalid_real_name'}",
             "—",
             "한글 2~10자 정규식"),
            ("T-505", "real_name 1자 오류",
             "T-501 조건",
             "real_name='김' 전달",
             "{success:false, error:'invalid_real_name'}",
             "—",
             "최소 2자"),
            ("T-506", "account_number 형식 오류 (9자리)",
             "T-501 조건",
             "account_number='123456789'",
             "{success:false, error:'invalid_account_number'}",
             "—",
             "10~14자리 숫자"),
            ("T-507", "account_number 형식 오류 (문자 포함)",
             "T-501 조건",
             "account_number='1234-5678-9012'",
             "{success:false, error:'invalid_account_number'}",
             "—",
             "숫자만 허용"),
            ("T-508", "이미 제출된 당첨자 재제출 시도",
             "account_submitted_at IS NOT NULL",
             "동일 요청 재전송",
             "{success:false, error:'already_submitted'}",
             "SELECT account_submitted_at FROM draw_winners WHERE id='<id>';",
             ""),
            ("T-509", "당첨자 아닌 유저가 계좌 제출",
             "draw_winners에 해당 user+draw 행 없음",
             "submit-bank-account 호출",
             "{success:false, error:'winner_not_found'}",
             "—",
             ""),
            ("T-510", "마감(7일) 초과 후 제출 시도",
             "draws.draw_date + 7d < now()",
             "submit-bank-account 호출",
             "{success:false, error:'deadline_expired'}",
             "SELECT draw_date, draw_date + interval '7 days' AS deadline FROM draws WHERE id='<draw_id>';",
             ""),
            ("T-511", "payment_status='paid'인 당첨자 제출 시도",
             "payment_status='paid'",
             "submit-bank-account 호출",
             "{success:false, error:'winner_not_found'} (pending 조건 미충족)",
             "SELECT payment_status FROM draw_winners WHERE draw_id='<id>' AND user_id='<uid>';",
             ""),
            ("T-512", "미인증(JWT 없음) 요청",
             "—",
             "Authorization 헤더 없이 POST",
             "401 Unauthorized",
             "—",
             ""),
            ("T-513", "get_pending_win RPC — 계좌 제출 전 조회",
             "T-501 이전 상태",
             "SELECT * FROM public.get_pending_win() (authenticated 유저)",
             "draw_id, prize_rank, deadline_at, payment_status='pending', account_submitted_at=null 반환",
             "SELECT * FROM public.get_pending_win();",
             "real_name 등 계좌 정보 포함 안 됨"),
            ("T-514", "get_pending_win RPC — 계좌 제출 후 조회",
             "account_submitted_at IS NOT NULL",
             "SELECT * FROM public.get_pending_win() (authenticated 유저)",
             "0행 반환 (account_submitted_at 있으면 제외)",
             "SELECT * FROM public.get_pending_win();",
             ""),
        ]
    },
    {
        "title": "G6_어드민관리",
        "color": COLORS["group6"],
        "desc": "어드민 당첨자 관리",
        "rows": [
            ("T-601", "어드민 당첨자 목록 조회 (1~3등)",
             "draw_winners 데이터 존재",
             "어드민 draws 페이지 접속, 해당 회차 선택",
             "prize_rank 1,2,3 당첨자 표시 (profiles.nickname 포함)",
             "SELECT dw.prize_rank, dw.user_id, p.nickname FROM draw_winners dw LEFT JOIN profiles p ON p.id=dw.user_id WHERE dw.draw_id='<id>' AND dw.prize_rank<=3;",
             "getWinners1to3 캐시 5분"),
            ("T-602", "등수별 요약 (WinnerSummary) 조회",
             "draw_winners, draw_prizes 존재",
             "어드민 페이지 summary 섹션 확인",
             "1~5등별 당첨자 수, per_winner_amount 표시",
             "SELECT dw.prize_rank, count(*), dp.amount/count(*) AS per_winner FROM draw_winners dw JOIN draw_prizes dp ON dp.draw_id=dw.draw_id AND dp.prize_rank=dw.prize_rank WHERE dw.draw_id='<id>' GROUP BY dw.prize_rank, dp.amount;",
             ""),
            ("T-603", "계좌 확인 토글 (false→true)",
             "draw_winners 행 존재",
             "AccountVerifyToggle 클릭 (false→true)",
             "draw_winners.account_verified=true로 변경",
             "SELECT account_verified FROM draw_winners WHERE id='<winner_id>';",
             "toggleAccountVerified"),
            ("T-604", "계좌 확인 토글 해제 (true→false)",
             "account_verified=true",
             "AccountVerifyToggle 클릭 (true→false)",
             "account_verified=false",
             "SELECT account_verified FROM draw_winners WHERE id='<winner_id>';",
             ""),
            ("T-605", "지급 상태 변경 — pending→paid",
             "draw_winners.payment_status='pending'",
             "PaymentStatusButton 클릭 → paid 선택",
             "payment_status='paid', paid_at=now()",
             "SELECT payment_status, paid_at FROM draw_winners WHERE id='<winner_id>';",
             ""),
            ("T-606", "지급 상태 변경 — paid→pending",
             "payment_status='paid'",
             "pending으로 변경",
             "payment_status='pending', paid_at=null",
             "SELECT payment_status, paid_at FROM draw_winners WHERE id='<winner_id>';",
             "paid_at null 처리 확인"),
            ("T-607", "지급 상태 변경 — pending→cancelled",
             "payment_status='pending'",
             "cancelled로 변경",
             "payment_status='cancelled'",
             "SELECT payment_status FROM draw_winners WHERE id='<winner_id>';",
             ""),
            ("T-608", "어드민 메모 저장",
             "draw_winners 행 존재",
             "AdminMemoInput에 메모 입력 후 저장",
             "admin_memo 업데이트",
             "SELECT admin_memo FROM draw_winners WHERE id='<winner_id>';",
             ""),
            ("T-609", "어드민 메모 빈 문자열 저장",
             "기존 메모 있음",
             "메모 비운 후 저장",
             "admin_memo='' 또는 null",
             "SELECT admin_memo FROM draw_winners WHERE id='<winner_id>';",
             ""),
            ("T-610", "당첨자 캐시 무효화 확인",
             "지급 상태 변경 후",
             "어드민 페이지 새로고침",
             "revalidateTag('draw-winners') 동작, 최신 상태 반영",
             "—",
             "revalidateTag 호출 확인"),
            ("T-611", "get_my_draw_winner RPC 조회",
             "당첨자 유저, status='drawn' 이후",
             "SELECT * FROM public.get_my_draw_winner('<draw_id>') (authenticated)",
             "prize_rank, match_count, payment_status, winner_comment 반환 (계좌 정보 미포함)",
             "SELECT * FROM public.get_my_draw_winner('<draw_id>');",
             ""),
            ("T-612", "get_draw_public_stats 조회",
             "status='completed' 회차, draw_winners 존재",
             "SELECT public.get_draw_public_stats('<draw_id>')",
             "rank_summary(등수별 당첨자 수, 1인당 당첨금) + winner_cards(익명화 이름, comment 있는 당첨자)",
             "SELECT public.get_draw_public_stats('<draw_id>');",
             "anon 접근 가능"),
            ("T-613", "이름 익명화 2자 검증",
             "draw_winners.real_name='김이'",
             "get_draw_public_stats 조회",
             "display_name='김*'",
             "SELECT public._anonymize_name('김이');",
             "_anonymize_name 함수"),
            ("T-614", "이름 익명화 3자 검증",
             "draw_winners.real_name='홍길동'",
             "get_draw_public_stats 조회",
             "display_name='홍*동'",
             "SELECT public._anonymize_name('홍길동');",
             ""),
            ("T-615", "get_recent_winners_for_banner 조회",
             "최근 3회차 이상 completed/drawn, 1등 당첨자 존재",
             "SELECT * FROM public.get_recent_winners_for_banner()",
             "최대 3행, round_number, display_name(익명화), amount, winner_comment",
             "SELECT * FROM public.get_recent_winners_for_banner();",
             "anon 접근 가능"),
        ]
    },
    {
        "title": "G7_수동당첨자",
        "color": COLORS["group7"],
        "desc": "수동 당첨자 추가 (Manual Winner)",
        "rows": [
            ("T-701", "수동 당첨자 추가 — user_id 없이 (외부 인물)",
             "draws 존재",
             "AddWinnerDialog: user_id 미입력, prize_rank=1, real_name 입력 → 저장",
             "draw_winners 행 생성 (source='manual', user_id=null, manual_referral_code 자동 발급)",
             "SELECT source, user_id, manual_referral_code, prize_rank FROM draw_winners WHERE draw_id='<id>' AND source='manual';",
             "draw_prizes 없으면 자동 생성 (1등 1000만, 2/3등 각 100만)"),
            ("T-702", "수동 당첨자 추가 — user_id 포함",
             "기존 profiles.id 존재",
             "user_id 입력 포함",
             "draw_winners 행 생성, profiles 조인 가능",
             "SELECT dw.user_id, p.nickname FROM draw_winners dw JOIN profiles p ON p.id=dw.user_id WHERE dw.id='<winner_id>';",
             ""),
            ("T-703", "draw_prizes 자동 생성 확인",
             "해당 draw에 draw_prizes 없음",
             "T-701 실행",
             "draw_prizes에 1~3등 행 생성 (1등:1000만, 2등:100만, 3등:100만)",
             "SELECT prize_rank, amount FROM draw_prizes WHERE draw_id='<id>' ORDER BY prize_rank;",
             "ignoreDuplicates:true로 멱등"),
            ("T-704", "manual_referral_code 고유성 확인",
             "수동 당첨자 2명 추가",
             "T-701 두 번 실행",
             "각각 다른 6자리 대문자+숫자 코드 발급",
             "SELECT manual_referral_code FROM draw_winners WHERE draw_id='<id>' AND source='manual';",
             "generateUniqueReferralCode, 10회 재시도"),
            ("T-705", "동일 draw + user_id 중복 수동 당첨자 차단",
             "동일 user_id의 수동 당첨자 이미 존재",
             "동일 draw_id + user_id로 재추가",
             "draw_winners_draw_user_unique 위반 에러",
             "SELECT count(*) FROM draw_winners WHERE draw_id='<id>' AND user_id='<uid>';",
             "partial unique index (user_id IS NOT NULL)"),
            ("T-706", "user_id=null 수동 당첨자 여러 명 허용",
             "user_id 없는 수동 당첨자",
             "user_id 미입력 수동 당첨자 2명 추가",
             "두 행 모두 생성 (null은 unique index 제외)",
             "SELECT count(*) FROM draw_winners WHERE draw_id='<id>' AND source='manual' AND user_id IS NULL;",
             ""),
            ("T-707", "수동 당첨자 삭제",
             "source='manual' 당첨자 존재",
             "DeleteWinnerButton 클릭 → 확인",
             "draw_winners 행 삭제",
             "SELECT count(*) FROM draw_winners WHERE id='<winner_id>';",
             "deleteManualWinner: source='manual' 재확인 후 삭제"),
            ("T-708", "자동 당첨자(auto) 삭제 시도 차단",
             "source='auto' 당첨자",
             "deleteManualWinner 직접 호출",
             "Error '자동 당첨자는 삭제할 수 없습니다'",
             "SELECT source FROM draw_winners WHERE id='<auto_winner_id>';",
             "source 체크 후 거부"),
        ]
    },
    {
        "title": "G8_만료처리",
        "color": COLORS["group8"],
        "desc": "만료 처리 — expire_draw_winners cron",
        "rows": [
            ("T-801", "7일 초과 미제출 당첨자 자동 만료",
             "draw.draw_date + 7d < now(), payment_status='pending', account_submitted_at=null",
             "SELECT public.expire_draw_winners(); 수동 실행",
             "해당 draw_winners.payment_status='cancelled', updated_at=now()",
             "SELECT payment_status, updated_at FROM draw_winners WHERE draw_id='<id>' AND payment_status='cancelled';",
             "cron '0 1 * * *' = 매일 10:00 KST"),
            ("T-802", "7일 이내 당첨자는 만료 제외",
             "draw_date + 7d >= now()",
             "expire_draw_winners() 실행",
             "payment_status='pending' 유지",
             "SELECT payment_status FROM draw_winners WHERE draw_id='<id>';",
             ""),
            ("T-803", "계좌 제출 완료 당첨자는 만료 제외",
             "account_submitted_at IS NOT NULL",
             "expire_draw_winners() 실행",
             "payment_status 변경 없음",
             "SELECT payment_status, account_submitted_at FROM draw_winners WHERE draw_id='<id>';",
             "account_submitted_at 있으면 제외"),
            ("T-804", "이미 paid 당첨자 만료 제외",
             "payment_status='paid'",
             "expire_draw_winners() 실행",
             "payment_status='paid' 유지",
             "SELECT payment_status FROM draw_winners WHERE id='<winner_id>';",
             "pending 조건"),
            ("T-805", "만료 후 계좌 제출 시도 차단",
             "payment_status='cancelled'",
             "submit-bank-account 호출",
             "{success:false, error:'winner_not_found'} (pending 조건 미충족)",
             "SELECT payment_status FROM draw_winners WHERE draw_id='<id>' AND user_id='<uid>';",
             ""),
            ("T-806", "만료 후 get_pending_win 조회",
             "payment_status='cancelled'",
             "SELECT * FROM public.get_pending_win() (authenticated)",
             "0행 반환 (payment_status='pending' 조건 미충족)",
             "SELECT * FROM public.get_pending_win();",
             ""),
            ("T-807", "pg_cron 잡 등록 확인",
             "DB 접근 가능",
             "SELECT jobname, schedule, command FROM cron.job WHERE jobname='expire-draw-winners-daily';",
             "schedule='0 1 * * *', command='select public.expire_draw_winners()'",
             "SELECT jobname, schedule, command FROM cron.job WHERE jobname='expire-draw-winners-daily';",
             ""),
        ]
    },
    {
        "title": "G9_초기화",
        "color": COLORS["group9"],
        "desc": "초기화 — TestControlPanel 리셋",
        "rows": [
            ("T-901", "초기화 — drawn 상태에서",
             "draws.status='drawn'",
             "TestControlPanel [초기화] 클릭 → 확인",
             "draw_winners 전체 삭제, draw_entries.status='entered'으로 롤백, draws.status='active', winning_numbers=null",
             "SELECT status, winning_numbers FROM draws WHERE id='<id>'; SELECT count(*) FROM draw_winners WHERE draw_id='<id>'; SELECT status FROM draw_entries WHERE draw_id='<id>' LIMIT 5;",
             ""),
            ("T-902", "초기화 — closed 상태에서",
             "draws.status='closed'",
             "[초기화] 클릭",
             "draws.status='active' 복원, draw_entries 롤백",
             "SELECT status FROM draws WHERE id='<id>';",
             "resetActive = status !== 'completed'"),
            ("T-903", "초기화 — completed 상태에서 불가",
             "draws.status='completed'",
             "[초기화] 버튼 확인",
             "버튼 disabled (resetActive=false)",
             "—",
             "completed는 초기화 불가"),
            ("T-904", "draw_entries 상태 롤백 범위 확인",
             "draw_entries에 won, lost, entered 혼재",
             "[초기화] 실행",
             "won/lost만 entered로 롤백 (entered는 그대로)",
             "SELECT status, count(*) FROM draw_entries WHERE draw_id='<id>' GROUP BY status;",
             ".in('status',['won','lost']) 조건"),
            ("T-905", "초기화 후 재응모 가능 확인",
             "T-901 실행 후",
             "enter-draw 호출",
             "정상 응모 처리 (status='active')",
             "SELECT status FROM draw_entries WHERE draw_id='<id>' ORDER BY entered_at DESC LIMIT 1;",
             ""),
        ]
    },
    {
        "title": "GE_예외엣지케이스",
        "color": COLORS["groupE"],
        "desc": "예외 / 엣지 케이스",
        "rows": [
            ("T-E01", "동시 응모 — Race Condition",
             "티켓 잔액=1, 두 요청 동시 전송",
             "동시 POST x2",
             "첫 번째만 성공, 두 번째 'insufficient_tickets' (FOR UPDATE 행 잠금)",
             "SELECT total, spent FROM ticket_balances WHERE user_id='<uid>';",
             "FOR UPDATE로 race 방지"),
            ("T-E02", "같은 번호 조합 여러 장 응모 허용",
             "active 회차",
             "동일 lottery_numbers 2세트 전달",
             "두 draw_entries 행 모두 생성 (중복 번호 조합 허용)",
             "SELECT lottery_numbers FROM draw_entries WHERE draw_id='<id>' AND user_id='<uid>';",
             "UNIQUE 제약 없음"),
            ("T-E03", "동일 draw_entry_id로 draw_winners 중복 INSERT",
             "수동으로 동일 entry_id로 insert",
             "INSERT 시도",
             "ON CONFLICT DO NOTHING (draw_winners_entry_unique)",
             "SELECT count(*) FROM draw_winners WHERE draw_entry_id='<entry_id>';",
             "partial unique index"),
            ("T-E04", "한 draw에서 동일 user_id 자동+수동 당첨 중복 방지",
             "auto 당첨자 user_id=X 존재",
             "동일 user_id로 수동 당첨자 추가 시도",
             "draw_winners_draw_user_unique 위반 에러",
             "SELECT source FROM draw_winners WHERE draw_id='<id>' AND user_id='<uid>';",
             "partial unique (user_id IS NOT NULL)"),
            ("T-E05", "close-draw 호출 시 active 없는 경우",
             "모든 회차 closed/completed",
             "close-draw POST",
             "{closed:0, upcomingCreated:0} 정상 반환",
             "SELECT status FROM draws ORDER BY round_number DESC LIMIT 5;",
             ""),
            ("T-E06", "publish-and-open 호출 시 drawn 없는 경우",
             "drawn 회차 없음",
             "publish-and-open POST",
             "{completed:0, pushed:[], activated:N}",
             "SELECT status FROM draws WHERE status='upcoming';",
             "upcoming→active는 진행"),
            ("T-E07", "publish-and-open — upcoming 없는 폴백",
             "upcoming 없고 drawn 있음",
             "publish-and-open POST",
             "신규 active 회차 생성, {newDraw:{id, round_number}}",
             "SELECT round_number, status FROM draws ORDER BY round_number DESC LIMIT 2;",
             "폴백 로직"),
            ("T-E08", "당첨자 유저 계정 삭제 시 draw_winners 처리",
             "draw_winners.user_id=uid 존재",
             "profiles 삭제",
             "draw_winners.user_id=null (on delete set null)",
             "SELECT user_id FROM draw_winners WHERE id='<winner_id>';",
             "FK on delete set null"),
            ("T-E09", "draws 삭제 시 draw_winners cascade 삭제",
             "draws.id 참조 draw_winners 존재",
             "draws 행 삭제",
             "draw_winners도 cascade 삭제",
             "SELECT count(*) FROM draw_winners WHERE draw_id='<deleted_id>';",
             "on delete cascade"),
            ("T-E10", "prize_rank 범위 오류 (0 또는 6)",
             "—",
             "draw_winners INSERT prize_rank=6",
             "CHECK (prize_rank BETWEEN 1 AND 5) 위반",
             "—",
             "DB 레벨 제약"),
            ("T-E11", "winning_numbers 포맷 오류 (5개 번호)",
             "—",
             "draws UPDATE winning_numbers='{1,2,3,4,5}'",
             "draws_winning_numbers_format CHECK 위반",
             "—",
             "validate_winning_numbers 함수"),
            ("T-E12", "bonus_number 범위 오류 (0 또는 46)",
             "—",
             "draws UPDATE bonus_number=0",
             "draws_bonus_number_range CHECK 위반",
             "—",
             ""),
            ("T-E13", "get_pending_win — 비인증 유저 호출",
             "—",
             "anon으로 get_pending_win() 호출",
             "EXCEPTION 'Not authenticated' 또는 403",
             "—",
             "revoke from anon"),
            ("T-E14", "draw_winners RLS — authenticated 유저 직접 SELECT",
             "authenticated JWT로 draw_winners 직접 SELECT",
             "supabase-js로 직접 쿼리",
             "0행 반환 또는 권한 오류 (RLS: service_role 전용)",
             "—",
             "RPC만 통해서 조회 가능"),
            ("T-E15", "winner_comment 51자 초과 — DB 레벨",
             "—",
             "draw_winners UPDATE winner_comment=51자",
             "CHECK (char_length(winner_comment) <= 50) 위반",
             "—",
             ""),
            ("T-E16", "같은 prize_rank로 draw_prizes 중복 등록",
             "draw_prizes에 (draw_id, prize_rank=1) 존재",
             "동일 draw_id, prize_rank=1 INSERT",
             "draw_prizes_draw_rank_unique 위반",
             "SELECT count(*) FROM draw_prizes WHERE draw_id='<id>' AND prize_rank=1;",
             "unique index"),
            ("T-E17", "generateUniqueReferralCode 10회 실패",
             "코드 공간 고갈 (시뮬레이션)",
             "10회 재시도 강제 실패",
             "Error '초대코드 생성에 실패했습니다'",
             "—",
             "10회 재시도 후 throw"),
            ("T-E18", "어드민 새로고침 캐시 — draws-list 5분 TTL",
             "draws 변경 후 즉시 조회",
             "revalidateTag 없이 새로고침",
             "unstable_cache 300초 TTL, 변경 미반영 가능",
             "—",
             "명시적 revalidateTag 호출 필요"),
            ("T-E19", "TestControlPanel 프로덕션 경고 배너",
             "NEXT_PUBLIC_APP_ENV='production'",
             "어드민 페이지 접속",
             "'⚠️ PRODUCTION' 빨간 배너 + 경고 문구 표시",
             "—",
             "ENV_BADGE.warn=true"),
            ("T-E20", "발표 완료(completed) 회차 재발표 시도",
             "draws.status='completed'",
             "Step3 버튼 클릭 시도",
             "버튼 disabled (step3Active=false)",
             "—",
             ""),
        ]
    },
]

# ══════════════════════════════════════════════════════════
# 각 그룹별 시트 생성
# ══════════════════════════════════════════════════════════
for group in GROUPS:
    ws = wb.create_sheet(title=group["title"])
    ws.sheet_view.showGridLines = False

    # 그룹 타이틀
    ws.merge_cells("A1:I1")
    ws["A1"] = group["desc"]
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=11, color="1E3A5F")
    ws["A1"].fill = PatternFill("solid", fgColor=group["color"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # 결과 범례 (오른쪽 - A1에 넣되 우측 정렬)
    ws["A1"].value = f"{group['desc']}        결과: ✅ PASS  ❌ FAIL  ⏭ SKIP"

    # 헤더
    apply_header(ws, 2, COLS, COLORS["header_bg"])
    ws.row_dimensions[2].height = 22

    # 데이터 행
    for i, row_data in enumerate(group["rows"]):
        r = 3 + i
        bg = COLORS["row_odd"] if i % 2 == 0 else COLORS["row_even"]
        full_row = list(row_data) + ["", ""]   # 결과, 메모 빈칸
        write_row(ws, r, full_row, bg)
        ws.row_dimensions[r].height = 55

    # 열 너비
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 결과 열 배경 (H열) — 조건부 서식 대신 기본 흰색
    # 테스터가 직접 색 입력

    ws.freeze_panes = "B3"

# ══════════════════════════════════════════════════════════
# 저장
# ══════════════════════════════════════════════════════════
OUTPUT = "/Users/naklee/repo/lift/lucky-piggy-admin/docs/test-scenarios/lucky_piggy_draw_qa_시나리오.xlsx"
wb.save(OUTPUT)
print(f"✅ 저장 완료: {OUTPUT}")
print(f"   총 {sum(len(g['rows']) for g in GROUPS)}개 테스트 케이스")
